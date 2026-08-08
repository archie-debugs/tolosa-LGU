import json
import io
import csv
import re
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, Request, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session
from .. import models, schemas
from ..database import get_db
from ..core import UPLOAD_DIR, record_audit_log
import hashlib
from sqlalchemy.exc import IntegrityError
import os
import secrets
from fastapi.responses import HTMLResponse

router = APIRouter(prefix="/documents", tags=["documents"])


def _next_tracking_number(db: Session) -> str:
    rows = db.query(models.Document.tracking_number).filter(models.Document.tracking_number.like("DOC-%")).all()
    used_numbers = set()
    for (tracking_number,) in rows:
        if not tracking_number or not tracking_number.startswith("DOC-"):
            continue
        suffix = tracking_number.replace("DOC-", "", 1)
        if suffix.isdigit():
            used_numbers.add(int(suffix))

    candidate = 1
    while candidate in used_numbers:
        candidate += 1
    return f"DOC-{candidate}"


def _serialize_document(doc: models.Document):
    payload = {
        "id": doc.id,
        "tracking_number": doc.tracking_number,
        "title": doc.title,
        "description": doc.description,
        "document_type": doc.document_type,
        "category": doc.category,
        "originating_office": doc.originating_office,
        "current_office": doc.current_office,
        "assigned_to": doc.assigned_to,
        "status": doc.status,
        "priority": doc.priority,
        "remarks": doc.remarks,
        "author": doc.author,
        "session": doc.session,
        "date_registered": doc.date_registered,
        "attachment_name": doc.attachment_name,
        "qr_code_value": doc.qr_code_value,
        "routing_history": [],
        "created_by": doc.created_by,
        "created_by_id": getattr(doc, "created_by_id", None),
        "document_type_id": getattr(doc, "document_type_id", None),
        "category_id": getattr(doc, "category_id", None),
        "originating_office_id": getattr(doc, "originating_office_id", None),
        "current_office_id": getattr(doc, "current_office_id", None),
        "archived": bool(doc.archived),
        "created_at": doc.created_at,
        "updated_at": doc.updated_at,
    }
    if doc.routing_history:
        try:
            payload["routing_history"] = json.loads(doc.routing_history)
        except Exception:
            payload["routing_history"] = []
    # include normalized history rows
    try:
        hist_rows = db_session = None
        from ..database import SessionLocal
        db_session = SessionLocal()
        rows = (
            db_session.query(models.DocumentHistory)
            .filter(models.DocumentHistory.document_id == doc.id)
            .order_by(models.DocumentHistory.created_at.asc())
            .all()
        )
        payload["history_rows"] = [
            {
                "id": r.id,
                "action": r.action,
                "actor": r.actor,
                "from_office": r.from_office,
                "to_office": r.to_office,
                "notes": r.notes,
                "created_at": r.created_at,
            }
            for r in rows
        ]
    except Exception:
        payload["history_rows"] = []
    finally:
        try:
            if db_session:
                db_session.close()
        except Exception:
            pass
    # include attachments metadata
    try:
        db_session = SessionLocal()
        atts = db_session.query(models.Attachment).filter(models.Attachment.document_id == doc.id).all()
        payload["attachments"] = [
            {"id": a.id, "original_filename": a.original_filename, "stored_path": a.stored_path, "mime_type": a.mime_type, "size": a.size, "checksum": a.checksum} for a in atts
        ]
    except Exception:
        payload["attachments"] = []
    finally:
        try:
            if db_session:
                db_session.close()
        except Exception:
            pass
    return payload


def _get_or_create_by_name(db: Session, Model, name: str):
    if not name:
        return None
    name = name.strip()
    existing = db.query(Model).filter(Model.name == name).first()
    if existing:
        return existing.id
    try:
        obj = Model(name=name)
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj.id
    except IntegrityError:
        db.rollback()
        existing = db.query(Model).filter(Model.name == name).first()
        return existing.id if existing else None


def _normalize_tracking_query(search: str | None) -> str | None:
    if search is None:
        return None
    text = str(search).strip()
    if not text:
        return None
    match = re.fullmatch(r"(?i)(?:doc[-\s]*)?0*([1-9][0-9]*|0)", text)
    if match:
        return match.group(1)
    return None


@router.get("", response_model=list[schemas.DocumentResponse])
def list_documents(
    search: str | None = Query(default=None),
    status: str | None = None,
    document_type: str | None = None,
    category: str | None = None,
    current_office: str | None = None,
    year: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
):
        query = db.query(models.Document).filter(models.Document.archived.is_(False))
        if search:
            normalized_tracking = _normalize_tracking_query(search)
            if normalized_tracking is not None and re.fullmatch(r"(?i)(?:doc[-\s]*)?0*([1-9][0-9]*|0)", str(search).strip()):
                canonical_tracking = f"DOC-{int(normalized_tracking)}"
                query = query.filter(
                    or_(
                        models.Document.tracking_number == canonical_tracking,
                        models.Document.tracking_number == str(search).strip().upper(),
                    )
                )
            else:
                like = f"%{search.lower()}%"
                query = query.filter(
                    or_(
                        models.Document.tracking_number.ilike(like),
                        models.Document.title.ilike(like),
                        models.Document.description.ilike(like),
                        models.Document.document_type.ilike(like),
                        models.Document.category.ilike(like),
                        models.Document.originating_office.ilike(like),
                        models.Document.current_office.ilike(like),
                        models.Document.assigned_to.ilike(like),
                        models.Document.status.ilike(like),
                        models.Document.remarks.ilike(like),
                        models.Document.author.ilike(like),
                        models.Document.session.ilike(like),
                        models.Document.created_by.ilike(like),
                    )
                )
        if status:
            query = query.filter(models.Document.status == status)
        if document_type:
            query = query.filter(models.Document.document_type == document_type)
        if category:
            query = query.filter(models.Document.category == category)
        if current_office:
            query = query.filter(models.Document.current_office == current_office)
        if year:
            year_text = str(year).strip()
            if year_text.isdigit() and len(year_text) == 4:
                start_year = datetime(int(year_text), 1, 1, tzinfo=timezone.utc)
                end_year = datetime(int(year_text) + 1, 1, 1, tzinfo=timezone.utc)
                query = query.filter(models.Document.created_at >= start_year, models.Document.created_at < end_year)
        if start_date:
            query = query.filter(models.Document.created_at >= start_date)
        if end_date:
            query = query.filter(models.Document.created_at <= end_date)

        docs = query.order_by(models.Document.created_at.desc()).all()
        return [_serialize_document(doc) for doc in docs]




@router.get("/{document_id:int}", response_model=schemas.DocumentResponse)
def get_document(document_id: int, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return _serialize_document(doc)


@router.post("", response_model=schemas.DocumentResponse, status_code=status.HTTP_201_CREATED)
def create_document(
    tracking_number: str | None = Form(None),
    title: str = Form(...),
    description: str | None = Form(None),
    document_type: str | None = Form(None),
    category: str | None = Form(None),
    originating_office: str | None = Form(None),
    current_office: str | None = Form(None),
    assigned_to: str | None = Form(None),
    status_field: str | None = Form("Pending"),
    priority: str | None = Form("Medium"),
    remarks: str | None = Form(None),
    author: str | None = Form(None),
    session: str | None = Form(None),
    date_registered: str | None = Form(None),
    qr_code_value: str | None = Form(None),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
):
    # basic validation
    if not title or not title.strip():
        raise HTTPException(status_code=400, detail="Title is required")

    tracking = (tracking_number or "").strip()
    if tracking:
        existing = db.query(models.Document).filter(models.Document.tracking_number == tracking).first()
        if existing:
            raise HTTPException(status_code=409, detail="Tracking number already exists")

    data = {
        "tracking_number": tracking or None,
        "title": title.strip(),
        "description": (description or None),
        "document_type": (document_type or None),
        "category": (category or None),
        "originating_office": (originating_office or None),
        "current_office": (current_office or None),
        "assigned_to": (assigned_to or None),
        "status": (status_field or "Pending"),
        "priority": (priority or "Medium"),
        "remarks": (remarks or None),
        "author": (author or None),
        "session": (session or None),
        "date_registered": (date_registered or None),
        "attachment_name": None,
        "qr_code_value": (qr_code_value or None),
        "routing_history": "[]",
        "created_by": None,
    }

    # ensure tracking number is present (DB requires non-null)
    if not data.get("tracking_number"):
        data["tracking_number"] = _next_tracking_number(db)
    if not data.get("qr_code_value"):
        data["qr_code_value"] = data["tracking_number"]

    # handle file upload (store into UPLOAD_DIR)
    if file is not None:
        try:
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            content = file.file.read()
            checksum = hashlib.sha256(content).hexdigest()
            safe_name = f"{int(datetime.utcnow().timestamp())}_{secrets.token_hex(8)}_{os.path.basename(file.filename)}"
            dest_path = os.path.join(UPLOAD_DIR, safe_name)
            with open(dest_path, "wb") as fh:
                fh.write(content)
            data["attachment_name"] = safe_name
            # persist attachment metadata after doc is created
            attachment_meta = {"original_filename": file.filename, "stored_path": dest_path, "mime_type": file.content_type or mimetypes.guess_type(file.filename)[0], "size": len(content), "checksum": checksum}
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Failed to save uploaded file: {exc}")

    # (removed debug artifact write)
    # create the document row (legacy text fields kept for compatibility)
    doc = models.Document(**{k: v for k, v in data.items() if v is not None})
    db.add(doc)
    db.commit()
    db.refresh(doc)

    # populate normalized FK columns if provided
    if document_type:
        doc.document_type_id = _get_or_create_by_name(db, models.DocumentType, document_type)
    if category:
        doc.category_id = _get_or_create_by_name(db, models.Category, category)
    if originating_office:
        doc.originating_office_id = _get_or_create_by_name(db, models.Office, originating_office)
    if current_office:
        doc.current_office_id = _get_or_create_by_name(db, models.Office, current_office)
    if assigned_to:
        # do not auto-create users here; leave created_by/assigned_to strings until manual mapping is done
        pass
    db.commit()
    db.refresh(doc)

    # store attachment metadata row if upload occurred
    try:
        if file is not None and 'attachment_meta' in locals():
            att = models.Attachment(
                document_id=doc.id,
                original_filename=attachment_meta['original_filename'],
                stored_path=attachment_meta['stored_path'],
                mime_type=attachment_meta.get('mime_type'),
                size=attachment_meta.get('size'),
                checksum=attachment_meta.get('checksum'),
            )
            db.add(att)
            db.commit()
            db.refresh(att)
    except Exception:
        db.rollback()

    # finalize routing history and ensure qr/tracking values persisted
    if not doc.qr_code_value:
        doc.qr_code_value = doc.tracking_number or f"DOC-{doc.id}"
    if not doc.routing_history:
        doc.routing_history = "[]"
    db.commit()
    db.refresh(doc)
    return _serialize_document(doc)


@router.put("/{document_id:int}", response_model=schemas.DocumentResponse)
def update_document(document_id: int, payload: schemas.DocumentUpdate, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    if payload.tracking_number is not None and payload.tracking_number != doc.tracking_number:
        existing = db.query(models.Document).filter(models.Document.tracking_number == payload.tracking_number).first()
        if existing:
            raise HTTPException(status_code=409, detail="Tracking number already exists")

    data = payload.model_dump(exclude_unset=True)
    # apply simple fields
    for field, value in data.items():
        if field in ("document_type", "category", "originating_office", "current_office"):
            # handled below to populate normalized FK columns
            continue
        setattr(doc, field, value)

    # handle normalized lookups
    if data.get("document_type") is not None:
        doc.document_type = data.get("document_type")
        doc.document_type_id = _get_or_create_by_name(db, models.DocumentType, data.get("document_type"))
    if data.get("category") is not None:
        doc.category = data.get("category")
        doc.category_id = _get_or_create_by_name(db, models.Category, data.get("category"))
    if data.get("originating_office") is not None:
        doc.originating_office = data.get("originating_office")
        doc.originating_office_id = _get_or_create_by_name(db, models.Office, data.get("originating_office"))
    if data.get("current_office") is not None:
        doc.current_office = data.get("current_office")
        doc.current_office_id = _get_or_create_by_name(db, models.Office, data.get("current_office"))

    db.commit()
    db.refresh(doc)
    return _serialize_document(doc)


@router.post("/{document_id:int}/route", response_model=schemas.DocumentResponse)
def route_document(document_id: int, payload: dict, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    destination_office = (payload.get("destination_office") or "").strip()
    assigned_user = (payload.get("assigned_user") or "").strip()
    remarks = (payload.get("remarks") or "").strip()
    route_label = (payload.get("route") or "Routing").strip() or "Routing"
    previous_status = doc.status or "Pending"
    from_office = doc.current_office or doc.originating_office or "Unknown"
    history = []
    if doc.routing_history:
        try:
            history = json.loads(doc.routing_history)
        except Exception:
            history = []

    history.append(
        {
            "date": payload.get("date") or "",
            "time": payload.get("time") or "",
            "from": from_office,
            "to": destination_office or "Unassigned",
            "user": assigned_user or doc.assigned_to or "Unassigned",
            "remarks": remarks or "Routed",
            "status": payload.get("status") or "In Routing",
            "route": route_label,
        }
    )
    doc.current_office = destination_office or doc.current_office
    doc.assigned_to = assigned_user or doc.assigned_to
    doc.status = payload.get("status") or "In Routing"
    doc.routing_history = json.dumps(history)
    # also persist a normalized history row for querying
    try:
        history_row = models.DocumentHistory(
            document_id=doc.id,
            action=route_label or "Routing",
            actor=assigned_user or doc.assigned_to or None,
            from_office=from_office,
            to_office=destination_office or None,
            notes=remarks or None,
        )
        db.add(history_row)
    except Exception:
        # best-effort: don't fail the route if history row cannot be created
        pass

    db.commit()
    db.refresh(doc)
    return _serialize_document(doc)


@router.post("/{document_id:int}/qr", response_model=schemas.DocumentResponse)
def regenerate_qr(document_id: int, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    doc.qr_code_value = doc.tracking_number or f"DOC-{doc.id}"
    db.commit()
    db.refresh(doc)
    return _serialize_document(doc)


@router.delete("/{document_id:int}")
def delete_document(document_id: int, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    doc.archived = True
    db.commit()
    return {"message": "Document archived successfully"}


@router.get("/template")
def download_import_template():
    try:
        try:
            import openpyxl
        except Exception:
            raise HTTPException(status_code=500, detail="Server missing openpyxl for template generation")
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Tracking Number",
            "Title",
            "Document Type",
            "Category",
            "Description",
            "Origin Office",
            "Current Office",
            "Current Holder",
            "Status",
            "Author",
            "Session",
            "Date Registered",
            "Remarks",
        ]
        ws.append(headers)
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=documents_import_template.xlsx"},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


def _parse_rows_from_file(file: UploadFile) -> list[dict]:
    name = (file.filename or "").lower()
    content = file.file.read()
    rows = []
    if name.endswith(".csv") or name.endswith(".txt"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append(r)
    else:
        try:
            import openpyxl
        except Exception:
            raise HTTPException(status_code=500, detail="Server missing openpyxl for xlsx parsing")
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        header = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            if not header:
                continue
            data = {}
            empty = True
            for h, cell in zip(header, row):
                if cell is not None and str(cell).strip() != "":
                    empty = False
                data[h or ""] = str(cell).strip() if cell is not None else ""
            if empty:
                continue
            rows.append(data)
    return rows


# Legacy import preview endpoint removed: import-by-file workflow deprecated


def _ensure_upload_dirs():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


# Removed separate upload page — uploads are handled inline via the admin frontend FilePicker


# Legacy import endpoints removed: import-by-file and temp upload workflow deprecated
